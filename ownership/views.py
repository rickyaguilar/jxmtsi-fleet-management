from django.shortcuts import render,HttpResponseRedirect,HttpResponse
from django.contrib.messages.views import SuccessMessageMixin
from django.views import generic
from openpyxl import Workbook
from django.urls import reverse_lazy
from .models import (
   	Ownership,
)
from masterlist.models import EmployeeMasterlist,VehicleMasterList
from . forms import (
    ownershipForm
)
from django.views.generic import (
                				CreateView,
                				ListView,
                				UpdateView,
                				DetailView,
                				)
from bootstrap_modal_forms.generic import (
                                           BSModalDeleteView
                                           )

class ownershipListView(ListView):
	model = Ownership
	template_name = 'transfer_list.html'

def ownershipcreate(request):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    e_list = EmployeeMasterlist.objects.all()
    v_list = VehicleMasterList.objects.all()
    return render(request, 'transfer_new.html',{'Title':'Ownership - Transfer of Ownership', 'e_list':e_list,'v_list':v_list})


def ownership_submit(request):
	if request.method == 'POST':
		date_application = request.POST.get('date_application')
		req_employee_id = request.POST.get('req_employee_id')
		req_Lname = request.POST.get('req_Lname')
		req_Fname = request.POST.get('req_Fname')
		req_band = request.POST.get('req_band')
		req_cost = request.POST.get('req_cost')
		req_title = request.POST.get('req_title')
		plate_no = request.POST.get('plate_no')
		cond_sticker = request.POST.get('cond_sticker')
		vehicle_model = request.POST.get('vehicle_model')
		vehicle_brand = request.POST.get('vehicle_brand')
		vehicle_make = request.POST.get('vehicle_make')
		Vendor = request.POST.get('Vendor')
		Other_Vendor_Name = request.POST.get('Other_Vendor_Name')
		v_employee_id = request.POST.get('v_employee_id')
		v_fname = request.POST.get('v_fname')
		v_lname = request.POST.get('v_lname')
		v_band = request.POST.get('v_band')
		TOO_Purpose = request.POST.get('TOO_Purpose')
		transfer_fee = request.POST.get('transfer_fee')
		doc_date_completed = request.POST.get('doc_date_completed')
		deedofsale_date = request.POST.get('deedofsale_date')
		confirmation_status = request.POST.get('confirmation_status')
		emailed_to_casher = request.POST.get('emailed_to_casher')
		received_from_casher = request.POST.get('received_from_casher')
		deed_signed = request.POST.get('deed_signed')
		routed_to_jd = request.POST.get('routed_to_jd')
		approved_by_jd = request.POST.get('approved_by_jd')
		return_fleet_admin = request.POST.get('return_fleet_admin')
		forwarded_to_liason = request.POST.get('forwarded_to_liason')
		date_notarized = request.POST.get('date_notarized')
		endorosed_to_insurance = request.POST.get('endorosed_to_insurance')
		requested_for_pullout = request.POST.get('requested_for_pullout')
		# date_pulled = request.POST.get('date_pulled')
		# return_endorsementfleet = request.POST.get('return_endorsementfleet')
		forwarded_fleet_liason = request.POST.get('forwarded_fleet_liason')
		tmg_date_in = request.POST.get('tmg_date_in')
		tmg_location = request.POST.get('tmg_location')
		tmg_date_return = request.POST.get('tmg_date_return')
		lto_location = request.POST.get('lto_location')
		lto_date_in = request.POST.get('lto_date_in')
		lto_date_out = request.POST.get('lto_date_out')
		# lto_date_return = request.POST.get('lto_date_return')
		# date_docs_return = request.POST.get('date_docs_return')
		date_transfered_completed = request.POST.get('date_transfered_completed')
		date_comletion_vismin = request.POST.get('date_comletion_vismin')
		date_received_by = request.POST.get('received_by')
		

		saveto_own = Ownership(date_application = date_application,req_employee_id = req_employee_id,req_Fname = req_Fname,req_Lname = req_Lname,
			req_band = req_band,req_cost = req_cost,req_title = req_title,plate_no= plate_no,cond_sticker = cond_sticker,vehicle_model = vehicle_model,
			vehicle_brand = vehicle_brand,vehicle_make = vehicle_make,vendor = Vendor,vendor_name = Other_Vendor_Name,v_employee_id =  v_employee_id,
			v_fname = v_fname,v_lname = v_lname,v_band = v_band,purpose = TOO_Purpose,transfer_fee = transfer_fee,doc_date_completed = doc_date_completed,
			deedofsale_date = deedofsale_date,confirmation_status = confirmation_status,emailed_to_casher = emailed_to_casher,received_from_casher = received_from_casher,
			deed_signed = deed_signed,routed_to_jd =routed_to_jd ,approved_by_jd =approved_by_jd ,return_fleet_admin = return_fleet_admin,forwarded_to_liason = forwarded_to_liason,
			date_notarized = date_notarized,endorosed_to_insurance =endorosed_to_insurance ,requested_for_pullout = requested_for_pullout,
			forwarded_fleet_liason = forwarded_fleet_liason,tmg_date_in =tmg_date_in ,tmg_location = tmg_location,
			tmg_date_return = tmg_date_return,lto_date_in = lto_date_in,lto_date_out = lto_date_out, lto_location = lto_location,
			date_transfered_completed = date_transfered_completed,date_comletion_vismin = date_comletion_vismin, date_received_by = date_received_by)
		saveto_own.save()

		return HttpResponseRedirect('/Ownership/Ownership/')

class ownershipUpdate(SuccessMessageMixin, UpdateView):
	model = Ownership
	form_class = ownershipForm
	template_name = 'transfer_form.html'

	def get_success_message(self, cleaned_data):
		print(cleaned_data)
		return "Transfer of Ownership Updated Successfully!"
		
class ownershipDetails(DetailView):
	model = Ownership
	template_name = 'transfer_details.html'
	
class ownershipDeleteView(BSModalDeleteView):
    model = Ownership
    template_name = 'transfer_delete.html'
    success_message = 'Success: Item was deleted.'
    success_url = reverse_lazy('ownership_list')

def ownershipHistoryView(request):
    if request.method == "GET":
       obj = Ownership.history.all()

       return render(request, 'transfer_history.html', context={'object': obj})

def ownership_excel(request):
    own_queryset = Ownership.objects.all()   
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename= Transfer of Ownership.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Transfer of Ownership'

    columns = [
			'Date Application Received' ,
		    'Requisitioner Employee ID' ,
		    'Requisitioner First Name' ,
		    'Requisitioner Last Name' ,
		    'Requisitioner Band' ,
		    'Requisitioner Cost Center' ,
		    'Requisitioner Title' ,
		    'Plate No' ,
		    'Conduction Sticker' ,
		    'Vehicle Model' ,
		    'Vehicle Brand' ,
		    'Vehicle Make' ,
		    'Vendor' ,
		    'Other Vendor Name' ,
		    'Vendee Employee ID' ,
		    'Vendee First Name' ,
		    'Vendee Last Name' ,
		    'Vendee Band' ,
		    'Purpose' ,
		    'Transfer Fee' ,
		    'Document Completed Date' 
		    'Date Created Deed of Sales' ,
		    'Confirmation Status' ,
		    'Date OR Emailed to Cashier' ,
		    'Date OR Received from Cashier' ,
		    'Signed Deed Of Sale Completed Date' ,
		    'Date Routed to JD/ECS' ,
		    'Date Approved by JD/ECS' ,
		    'Date Return to Fleet Admin Assistant' ,
		    'Date Forwarded to Liason Officer' ,
		    'Date Notarized' ,
		    'Date Endorsed to Insurance Company' ,
		    'Date Received Pull-Out of OR/CR' ,
		    # 'Date Pulled OR/CR' ,
		    # 'Date Return Endorsed to Fleet Admin' ,
		    'Date Forwarded to Fleet Liason for TMG' ,
		    'TMG Date Input' ,
		    'TMG Location' ,
		    'TMG Date Return to Fleet Liason' ,
		    'LTO Date In' ,
		    'LTO Date Out' ,
		    'LTO Location' ,
		    # 'LTO Date Return' ,
		    # 'Date Documents Return' ,
		    'Date Transfer Completed' ,
		    'Date Transfer Completed for Visayas/Mindanao' ,
		    'Date Received By',
		    'SLA' ,
		    'Date Initiated' ,
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for own in own_queryset:
        row_num += 1
        row = [
			own.date_application ,
		    own.req_employee_id ,
		    own.req_Fname ,
		    own.req_Lname ,
		    own.req_band ,
		    own.req_cost ,
		    own.req_title ,
		    own.plate_no ,
		    own.cond_sticker ,
		    own.vehicle_model ,
		    own.vehicle_brand ,
		    own.vehicle_make ,
		    own.vendor ,
		    own.vendor_name ,
		    own.v_employee_id ,
		    own.v_fname ,
		    own.v_lname ,
		    own.v_band ,
		    own.purpose ,
		    own.transfer_fee ,
		    own.doc_date_completed ,
		    own.deedofsale_date ,
		    own.confirmation_status ,
		    own.emailed_to_casher ,
		    own.received_from_casher ,
		    own.deed_signed ,
		    own.routed_to_jd ,
		    own.approved_by_jd ,
		    own.return_fleet_admin ,
		    own.forwarded_to_liason ,
		    own.date_notarized ,
		    own.endorosed_to_insurance ,
		    own.requested_for_pullout ,
		    # own.date_pulled ,
		    # own.return_endorsementfleet ,
		    own.forwarded_fleet_liason ,
		    own.tmg_date_in ,
		    own.tmg_location ,
		    own.tmg_date_return ,
		    own.lto_date_in ,
		    own.lto_date_out ,
		    own.lto_location,
		    # own.lto_date_return ,
		    #own.date_docs_return ,
		    own.date_transfered_completed ,
		    own.date_comletion_vismin ,
		    own.date_received_by ,
		    own.TOO_SLA ,
		    own.date_initiated ,
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response



