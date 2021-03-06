from django.shortcuts import render,HttpResponseRedirect,HttpResponse
from django.contrib.messages.views import SuccessMessageMixin
from django.views import generic
from openpyxl import Workbook
from django.urls import reverse_lazy
from .models import (
   	Fata_monitoring,
)
from django.db.models import Q
from . forms import (
    FATAmonitoringForm,
    reg_updateForm
)
from masterlist.models import EmployeeMasterlist,VehicleMasterList
from django.views.generic import (
                				CreateView,
                				ListView,
                				UpdateView,
                				DetailView,
                				)
from bootstrap_modal_forms.generic import (
                                           BSModalDeleteView
                                           )

class monitoringListView(ListView):
	model = Fata_monitoring
	template_name = 'fata_monitoringlist.html'

def monitoring_create(request):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    e_list = EmployeeMasterlist.objects.all()
    v_list = VehicleMasterList.objects.all()
    return render(request, 'fata_monitoring_new.html',{'Title':'Monitoring - Fata Monitoring', 'e_list':e_list,'v_list':v_list})

def monitoring_submit(request):
	if request.method == 'POST':
		Fata_no = request.POST.get('Fata_no')
		Date_transfer = request.POST.get('Date_transfer')
		Date_received = request.POST.get('Date_received')
		Plate_no = request.POST.get('Plate_no')
		v_make = request.POST.get('v_make')
		v_brand = request.POST.get('v_brand')
		Certificate_of_Reg = request.POST.get('Certificate_of_Reg')
		v_model = request.POST.get('v_model')
		Transferor_employee = request.POST.get('Transferor_employee')
		Transferor_Fname = request.POST.get('Transferor_Fname')
		Transferor_Lname = request.POST.get('Transferor_Lname')
		Recipient_Employee = request.POST.get('Recipient_Employee')
		Recipient_Fname = request.POST.get('Recipient_Fname')
		Recipient_Lname = request.POST.get('Recipient_Lname')
		Date_endorsed_Globe = request.POST.get('Date_endorsed_Globe')
		Date_endorsed_Innove = request.POST.get('Date_endorsed_Innove')
		Clearing_accountability = request.POST.get('Clearing_accountability')
		Globe_fixed_asset = request.POST.get('Globe_fixed_asset')
		Innove_fixed_asset = request.POST.get('Innove_fixed_asset')
		

		saveto_fata = Fata_monitoring(Fata_no=Fata_no ,Date_transfer=Date_transfer ,Date_received=Date_received ,Plate_no=Plate_no ,
			Vehicle_make=v_make ,Vehicle_brand=v_brand ,Certificate_of_Reg=Certificate_of_Reg ,Vehicle_model=v_model ,
			Transferor_employee=Transferor_employee ,Transferor_Fname=Transferor_Fname ,Transferor_Lname=Transferor_Lname ,
			Recipient_Employee=Recipient_Employee ,Recipient_Fname=Recipient_Fname ,Recipient_Lname=Recipient_Lname ,Date_endorsed_Globe=Date_endorsed_Globe ,
			Date_endorsed_Innove=Date_endorsed_Innove ,Clearing_accountability=Clearing_accountability ,Globe_fixed_asset=Globe_fixed_asset ,Innove_fixed_asset=Innove_fixed_asset)
		saveto_fata.save()

		return HttpResponseRedirect('/Monitoring/Monitoring/')
		
class monitoringCreateView(SuccessMessageMixin, CreateView):
	def dispatch(self, *args, **kwargs):
	    return super().dispatch(*args, **kwargs)
	template_name = 'fata_monitoring.html'
	form_class = FATAmonitoringForm

	def get_success_message(self, cleaned_data):
		print(cleaned_data)
		return "FATA Monitoring Has been Created!"

class monitoringUpdate(SuccessMessageMixin, UpdateView):
	model = Fata_monitoring
	form_class = FATAmonitoringForm
	template_name = 'fata_monitoring.html'

	def get_success_message(self, cleaned_data):
		print(cleaned_data)
		return "FATA Monitoring Updated Successfully!"
		
class monitoringDetails(DetailView):
	model = Fata_monitoring
	template_name = 'fata_monitoring_details.html'
	
class monitoringDeleteView(BSModalDeleteView):
    model = Fata_monitoring
    template_name = 'monitoring/fata_monitoring_delete.html'
    success_message = 'Success: Item was deleted.'
    success_url = reverse_lazy('Monitoring_list')

class regUpdate(SuccessMessageMixin, UpdateView):
	model = VehicleMasterList
	form_class = reg_updateForm
	template_name = 'regupdate.html'
	success_url = reverse_lazy('Monitoring_jan_reg')
	def get_success_message(self, cleaned_data):
		print(cleaned_data)
		return "Registrations Updated Successfully!"

# def regUpdate(request, pk):
# 	if request.method == 'POST':
# 		Last_Registration_Date = request.POST.get('last_reg')
# 		Smoke_Emission_Date = request.POST.get('smoke_date')
# 		COC_Date = request.POST.get('coc_date')

# 		Remarks = ""
# 		if Last_Registration_Date == "":
# 			Remarks = "Without Last Registrations Date"
# 		elif Smoke_Emission_Date == "":
# 			Remarks = "Without Smoke Emission Date"
# 		elif COC_Date == "":
# 			Remarks = "COC_Date"
# 		else:
# 			Remarks = "Complete"

# 		VehicleMasterList.objects.filter(id=pk).update(Last_Registration_Date=Last_Registration_Date,Smoke_Emission_Date=Smoke_Emission_Date,COC_Date=COC_Date,Remarks=Remarks)

# 		return HttpResponseRedirect('/Monitoring/Registration/January')

def monitoringHistoryView(request):
    if request.method == "GET":
       obj = Fata_monitoring.history.all()

       return render(request, 'fata_monitoring_history.html', context={'object': obj})

def janRegView(request):
	context = {
			'jan_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="JAN", Last_Registration_Date__isnull=True ),
			'jan_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="JAN", Smoke_Emission_Date__isnull=True),
			'jan_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="JAN", COC_Date__isnull=True)	
		}

	return render(request, 'regJan_monitoring.html', context)

def febRegView(request):
	context = {
			'feb_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="FEB", Last_Registration_Date__isnull=True),
			'feb_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="FEB", Smoke_Emission_Date__isnull=True),
			'feb_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="FEB", COC_Date__isnull=True)
		}

	return render(request, 'regFeb_monitoring.html', context)

def marRegView(request):
	context = {
			'mar_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="MAR",Last_Registration_Date__isnull=True),
			'mar_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="MAR",Smoke_Emission_Date__isnull=True),
			'mar_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="MAR",COC_Date__isnull=True)
		}

	return render(request, 'regMar_monitoring.html', context)

def aprRegView(request):
	context = {
			'apr_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="APR", Last_Registration_Date__isnull=True),
			'apr_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="APR", Smoke_Emission_Date__isnull=True),
			'apr_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="APR", COC_Date__isnull=True)

		}

	return render(request, 'regApr_monitoring.html', context)

def mayRegView(request):
	context = {
			'may_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="MAY", Last_Registration_Date__isnull=True),
			'may_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="MAY", Smoke_Emission_Date__isnull=True),
			'may_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="MAY", COC_Date__isnull=True)
		}

	return render(request, 'regMay_monitoring.html', context)

def junRegView(request):
	context = {
			'jun_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="JUN", Last_Registration_Date__isnull=True),
			'jun_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="JUN", Smoke_Emission_Date__isnull=True),
			'jun_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="JUN", COC_Date__isnull=True)
		}

	return render(request, 'regJun_monitoring.html', context)

def julRegView(request):
	context = {
			'jul_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="JUL", Last_Registration_Date__isnull=True),
			'jul_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="JUL", Smoke_Emission_Date__isnull=True),
			'jul_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="JUL", COC_Date__isnull=True)
		}
	return render(request, 'regJul_monitoring.html', context)

def augRegView(request):
	context = {
			'aug_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="AUG", Last_Registration_Date__isnull=True),
			'aug_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="AUG", Smoke_Emission_Date__isnull=True),
			'aug_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="AUG", COC_Date__isnull=True)
		}

	return render(request, 'regAug_monitoring.html', context)

def sepRegView(request):
	context = {
			'sep_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="SEP", Last_Registration_Date__isnull=True),
			'sep_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="SEP", Smoke_Emission_Date__isnull=True),
			'sep_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="SEP", COC_Date__isnull=True)
		}

	return render(request, 'regSep_monitoring.html', context)

def octRegView(request):
	context = {
			'oct_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="OCT", Last_Registration_Date__isnull=True),
			'oct_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="OCT", Smoke_Emission_Date__isnull=True),
			'oct_list': VehicleMasterList.objects.filter(REGISTRATION_MONTH__contains="OCT", COC_Date__isnull=True)
		}

	return render(request, 'regOct_monitoring.html', context)

def plateMonitoringView(request):
	context = {
			'plate_monitoring': VehicleMasterList.objects.filter(PLATE_NO__isnull=True)
		}

	return render(request, 'plate_monitoring.html', context)

def fata_excel(request):
    fata_queryset = Fata_monitoring.objects.all()   
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename= FATA Monitoring.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'FATA Monitoring'

    columns = [
				'FATA Number' ,
				'Date Transfer' ,
				'Date Received' ,
				'Plate No' ,
				'Vehicle Make' ,
				'Vehicle Brand' ,
				'Certificate Of Registrations Name' ,
				'Vehicle Model' ,
				'Transferor Employee' ,
				'Transferor First Name' ,
				'Transferor Last Name' ,
				'Recipient Employee' ,
				'Recipient Fist Name' ,
				'Recipient Last Name' ,
				'Date Endorsed Globe' ,
				'Date Endorsed Innove' ,
				'Clearing of Accountability' ,
				'Globe Fixed Asset Recepient' ,
				'Innove Fixed Asset Recepient' ,
				'Date Initiated' ,

    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for fata in fata_queryset:
        row_num += 1
        row = [
				fata.Fata_no ,
				fata.Date_transfer ,
				fata.Date_received ,
				fata.Plate_no ,
				fata.Vehicle_make ,
				fata.Vehicle_brand ,
				fata.Certificate_of_Reg ,
				fata.Vehicle_model ,
				fata.Transferor_employee ,
				fata.Transferor_Fname ,
				fata.Transferor_Lname ,
				fata.Recipient_Employee ,
				fata.Recipient_Fname ,
				fata.Recipient_Lname ,
				fata.Date_endorsed_Globe ,
				fata.Date_endorsed_Innove ,
				fata.Clearing_accountability ,
				fata.Globe_fixed_asset ,
				fata.Innove_fixed_asset ,
				fata.Date_initiated ,
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response


		
		