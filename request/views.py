from django.views import generic
from django.shortcuts import render,HttpResponseRedirect, get_list_or_404,HttpResponse
from django.contrib.messages.views import SuccessMessageMixin
from django.urls import reverse_lazy
from openpyxl import Workbook
from .models import (
		CarRentalRequest,
        Gas_card,
        service_vehicle,
        Vehicle_Repair,
)
from masterlist.models import EmployeeMasterlist,VehicleMasterList
from django.views.generic import (
     DetailView,
     ListView,
     CreateView,
     UpdateView,
 )
# from django.views.generic.edit import (
    # DeleteView,
# )
from . forms import (
    carrequestform,
    gascardform,
    serviceform,
    repairform,

    )

from bootstrap_modal_forms.generic import (
                                           BSModalDeleteView
                                           )


                      #####################################  
                    #########################################
                   #####.       Car Rental Request       #####
                    #########################################
                     ######################################



def requestCreate(request):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    emplist = EmployeeMasterlist.objects.all()
    return render(request, 'car_rental/carrequest_new.html',{'Title':'Car - Car Request', 'emplist':emplist})

def requestsubmit(request):
    if request.method == 'POST':
        emp_id = request.POST.get('emp_id')
        date_received = request.POST.get('date_received')
        fname = request.POST.get('fname')
        lname = request.POST.get('lname')
        cnumber = request.POST.get('cnumber')
        company = request.POST.get('company')
        band = request.POST.get('band')
        dept = request.POST.get('dept')
        cost = request.POST.get('cost')
        div = request.POST.get('div')
        loc = request.POST.get('loc')
        section = request.POST.get('section')
        designation = request.POST.get('designation')
        atd = request.POST.get('atd')
        vname = request.POST.get('vname')
        date = request.POST.get('date')
        up_to = request.POST.get('up_to')
        time = request.POST.get('time')
        del_place = request.POST.get('del_place')
        type_rental = request.POST.get('type_rental')
        costcenter = request.POST.get('costcenter')
        rent_period = request.POST.get('rent_period')
        destination = request.POST.get('destination')
        del_date = request.POST.get('del_date')
        end_user = request.POST.get('end_user')
        vehicle_type = request.POST.get('vehicle_type')
        plate_no = request.POST.get('plate_no')
        supervisor = request.POST.get('supervisor')
        cr_sla = request.POST.get('cr_sla')

        saveto_req = CarRentalRequest(A_Employee=emp_id, Date_received = date_received,Assignee_Fname = fname,Assignee_Lname = lname,Assignee_No = cnumber,Assignee_Company = company,
                Assignee_band = band,Assignee_Dept = dept,Assignee_Cost = cost,Assignee_Div = div,Assignee_Loc = loc,Assignee_Section = section,
                Assignee_Designation = designation,Assignee_ATD = atd,Vendor_name = vname,Date = date,Up_to = up_to,Time = time,Place_of_del = del_place,
                type_rental = type_rental,Cost_center = costcenter,Rental_period = rent_period,Destination = destination,Delivery_date = del_date,End_user = end_user,Type_of_vehicle = vehicle_type,
                Plate_no = plate_no,Immediate_supervisor =supervisor ,CR_SLA = cr_sla)
        saveto_req.save()

        return HttpResponseRedirect('/Request/Request/')

class requestListView(ListView):
	model = CarRentalRequest
	template_name = 'car_rental/carrequest_list.html'

class requestDetailView(DetailView):
	model = CarRentalRequest
	template_name = 'car_rental/carrequest_details.html'

class requestUpdateView(SuccessMessageMixin, UpdateView):
    model = CarRentalRequest
    form_class = carrequestform
    template_name = 'car_rental/carrequest_form.html'
    def get_success_message(self, cleaned_data):
    	print(cleaned_data)
    	return "Car Rental Request Updated Successfully!"

class requestDeleteView(BSModalDeleteView):
    model = CarRentalRequest
    template_name = 'car_rental/car_delete.html'
    success_message = 'Success: Report was deleted.'
    success_url = reverse_lazy('carrequest_list')

def requestHistoryView(request):
    if request.method == "GET":
       obj = CarRentalRequest.history.all()

       return render(request, 'car_rental/carrequest_history.html', context={'object': obj})

def car_request_excel(request):
    rental_queryset = CarRentalRequest.objects.all()   
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Car Rental Request.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Car Rental Request'

    columns = [
                'Assignee Employee' ,
                'Date Received' ,
                'Assignee First name' ,
                'Assignee Last name' ,
                'Assignee No' ,
                'Assignee Company' ,
                'Assignee Band' ,
                'Assignee Department' ,
                'Assignee Cost Center' ,
                'Assignee Division' ,
                'Assignee Loccation' ,
                'Assignee Section' ,
                'Assignee Designation' ,
                'Assignee ATD' ,
                'Vendor Name' ,
                'Date' ,
                'Up to' ,
                'Time' ,
                'Place of Delivery' ,
                'Type of Rental' ,
                'Cost Center' ,
                'Rental Reriod' ,
                'Destination' ,
                'Delivery Date' ,
                'End User' ,
                'Type of Vehicle' ,
                'Plate no' ,
                'Immediate Supervisor' ,
                'SLA' ,
                'Date Initiated' ,
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for car in rental_queryset:
        row_num += 1
        row = [
                car.A_Employee ,
                car.Date_received ,
                car.Assignee_Fname ,
                car.Assignee_Lname ,
                car.Assignee_No ,
                car.Assignee_Company ,
                car.Assignee_band ,
                car.Assignee_Dept ,
                car.Assignee_Cost ,
                car.Assignee_Div ,
                car.Assignee_Loc ,
                car.Assignee_Section ,
                car.Assignee_Designation ,
                car.Assignee_ATD ,
                car.Vendor_name ,
                car.Date ,
                car.Up_to ,
                car.Time ,
                car.Place_of_del ,
                car.type_rental ,
                car.Cost_center ,
                car.Rental_period ,
                car.Destination ,
                car.Delivery_date ,
                car.End_user ,
                car.Type_of_vehicle ,
                car.Plate_no ,
                car.Immediate_supervisor ,
                car.CR_SLA ,
                car.Date_initiated ,
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response


                      #####################################  
                    #########################################
                   #####.       Gas Card Request         #####
                    #########################################
                     ######################################


def gascreate(request):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    elist = EmployeeMasterlist.objects.all()
    vlist = VehicleMasterList.objects.all()
    return render(request, 'gas_card/gascard_new.html',{'Title':'Gas - Gas Card Request', 'elist':elist, 'vlist':vlist})
    # def get_success_message(self, cleaned_data):
    #    print(cleaned_data)
    #    return "New Car Rental Request Has been Created!"

def gassubmit(request):
    if request.method == 'POST':
        date_app = request.POST.get('date_app')
        app_type = request.POST.get('app_type')
        fleet_card = request.POST.get('fleet_card')
        fleet_card_type = request.POST.get('fleet_card_type')
        fuel_amount = request.POST.get('fuel_amount')
        fuel_quantity = request.POST.get('fuel_quantity')
        products_restriction = request.POST.get('products_restriction')
        req_emp_id = request.POST.get('req_emp_id')
        r_fname = request.POST.get('r_fname')
        r_lname = request.POST.get('r_lname')
        r_costcenter = request.POST.get('r_costcenter')
        r_title = request.POST.get('r_title')
        atd_no = request.POST.get('atd_no')
        temp_atd = request.POST.get('temp_atd')
        new_empId = request.POST.get('new_empId')
        new_fname = request.POST.get('new_fname')
        new_lname = request.POST.get('new_lname')
        new_costcenter = request.POST.get('new_costcenter')
        new_tempATD = request.POST.get('new_tempATD')
        new_assignee = request.POST.get('new_assignee')
        cost_code = request.POST.get('cost_code')
        gcr_cancel = request.POST.get('gcr_cancel')
        plate_no = request.POST.get('plate_no')
        c_sticker = request.POST.get('c_sticker')
        gcr_sla = request.POST.get('gcr_sla')
        v_brand = request.POST.get('v_brand')
        v_model = request.POST.get('v_model')
        v_make = request.POST.get('v_make')
        v_fueltype = request.POST.get('v_fueltype')
        new_plate_no = request.POST.get('new_plate_no')
        new_cs = request.POST.get('new_cs')
        new_model = request.POST.get('new_model')
        new_brand = request.POST.get('new_brand')
        new_make = request.POST.get('new_make')
        new_fueltype = request.POST.get('new_fueltype')
        approved_by = request.POST.get('approved_by')
        date_sumitted_app = request.POST.get('date_sumitted_app')
        date_recieved_fleet = request.POST.get('date_recieved_fleet')
        fleet_card_no = request.POST.get('fleet_card_no')
        fleet_card_release = request.POST.get('fleet_card_release')
        person_release_card = request.POST.get('person_release_card')
        gcr_sla = request.POST.get('gcr_sla')

        saveto_gas = Gas_card(date_received = date_app,application_type = app_type,fleet_provider= fleet_card,fleetcard_type =fleet_card_type ,
            fuel_limit_amount = fuel_amount,fuel_limit_quantity = fuel_quantity,products_restriction = products_restriction,req_employee = req_emp_id,
            req_fname = r_fname,req_lname = r_lname ,req_title = r_title,req_cost_center = r_costcenter, atd_no=atd_no, temporary_atd=temp_atd,
            new_emp_id=new_empId, new_emp_fname=new_fname, new_emp_lname=new_lname, new_emp_cost = new_costcenter,new_temp_atd = new_tempATD,
            new_assignee = new_assignee,cost_center_code = cost_code,cancellation = gcr_cancel,plate_no = plate_no,con_sticker = c_sticker,
            model_year = v_model,brand = v_brand,make = v_make,fuel_type = v_fueltype,new_plate_no = new_plate_no,new_cond_sticker = new_cs,
            new_model_year = new_model,new_vbrand = new_brand,new_vmake = new_make,new_vfuel_type = new_fueltype,approved_by = approved_by,
            date_summitted =date_sumitted_app,fleet_received = date_recieved_fleet,fleet_card_no = fleet_card_no,fleet_date_release = fleet_card_release,
            person_release = person_release_card,GCR_SLA = gcr_sla)
        saveto_gas.save()

        return HttpResponseRedirect('/Request/Gas/')
    
class gasListView(ListView):
    model = Gas_card
    template_name = 'gas_card/gascard_list.html'

class gasUpdateView(SuccessMessageMixin, UpdateView):
    model = Gas_card
    form_class = gascardform
    template_name = 'gas_card/gascard_form.html'
    
    def get_success_message(self, cleaned_data):
    	print(cleaned_data)
    	return "Gas Card Details Update Successfully!"

class gasDetailView(DetailView):
    model = Gas_card
    template_name = 'gas_card/gascard_details.html'

class gasDeleteView(BSModalDeleteView):
    model = Gas_card
    template_name = 'gas_card/gascard_delete.html'
    success_message = 'Success: Gas Gard Request was deleted.'
    success_url = reverse_lazy('gascard_list')

def gasHistoryView(request):
    if request.method == "GET":
       obj = Gas_card.history.all()

       return render(request, 'gas_card/gascard_history.html', context={'object': obj})

def gas_request_excel(request):
    rental_queryset = Gas_card.objects.all()   
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Gas Card Request.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Gas Card Request'

    columns = [
            'Date Received' ,
            'Application Type' ,
            'Fleet Provider' ,
            'Fleet Card Type' ,
            'Fuel Limit Amount' ,
            'Fuel Limit Quantity' ,
            'Products Restriction' ,
            'Employee' ,
            'First Name' ,
            'Last Name' ,
            'Title' ,
            'Cost Center' ,
            'ATD No' ,
            'Temporary ATD' ,
            'New Employee ID' ,
            'New employee First Name' ,
            'New employee Last Name' ,
            'New employee Cost Center' ,
            'New employee ATD' ,
            'New Assignee' ,
            'Cost Center Code' ,
            'Cancellation' ,
            'Plate No' ,
            'Conduction Sticker' ,
            'Model Year' ,
            'Brand' ,
            'Make' ,
            'Fuel Type' ,
            'New Plate No' ,
            'New Conduction Sticker' ,
            'New Model Year' ,
            'New Vehicle Brand' ,
            'New Vehicle Make' ,
            'New Vehicle fuel type' ,
            'Approved By' ,
            'Date Summitted' ,
            'Fleet Received' ,
            'Fleet Card No' ,
            'Fleet Date Release' ,
            'Person Release' ,
            'Date Initiated' ,
            'SLA' ,

    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for rental in rental_queryset:
        row_num += 1
        row = [
            rental.date_received ,
            rental.application_type ,
            rental.fleet_provider ,
            rental.fleetcard_type ,
            rental.fuel_limit_amount ,
            rental.fuel_limit_quantity ,
            rental.products_restriction ,
            rental.req_employee ,
            rental.req_fname ,
            rental.req_lname ,
            rental.req_title ,
            rental.req_cost_center ,
            rental.atd_no ,
            rental.temporary_atd ,
            rental.new_emp_id ,
            rental.new_emp_fname ,
            rental.new_emp_lname ,
            rental.new_emp_cost ,
            rental.new_temp_atd ,
            rental.new_assignee ,
            rental.cost_center_code ,
            rental.cancellation ,
            rental.plate_no ,
            rental.con_sticker ,
            rental.model_year ,
            rental.brand ,
            rental.make ,
            rental.fuel_type ,
            rental.new_plate_no ,
            rental.new_cond_sticker ,
            rental.new_model_year ,
            rental.new_vbrand ,
            rental.new_vmake ,
            rental.new_vfuel_type ,
            rental.approved_by ,
            rental.date_summitted ,
            rental.fleet_received ,
            rental.fleet_card_no ,
            rental.fleet_date_release ,
            rental.person_release ,
            rental.date_initiated ,
            rental.GCR_SLA ,
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response


                      #####################################  
                    #########################################
                   #####.     Service Vehicle Request    #####
                    #########################################
                     ######################################


class serviceListView(ListView):
    model = service_vehicle
    template_name = 'service_vehicle/service_list.html'

def serviceCreate(request):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    emplist = EmployeeMasterlist.objects.all()
    vlist = VehicleMasterList.objects.all()
    return render(request, 'service_vehicle/service_new.html',{'Title':'Car - Car Request', 'emplist':emplist,'vlist':vlist})

def servicesubmit(request):
    if request.method == 'POST':
        req_employee_id = request.POST.get('req_employee_id')
        request_date = request.POST.get('request_date')
        req_lname = request.POST.get('req_lname')
        req_fname = request.POST.get('req_fname')
        assignee_employee_id = request.POST.get('assignee_employee_id')
        Assignee_Group = request.POST.get('Assignee_Group')
        assignee_fname = request.POST.get('assignee_fname')
        assignee_lname = request.POST.get('assignee_lname')
        assignee_costcenter = request.POST.get('assignee_costcenter')
        assignee_section = request.POST.get('assignee_section')
        assignee_location = request.POST.get('assignee_location')
        assignee_atd = request.POST.get('assignee_atd')
        new_employee_id = request.POST.get('new_employee_id')
        new_employee_fname = request.POST.get('new_employee_fname')
        new_employee_lname = request.POST.get('new_employee_lname')
        new_employee_cost = request.POST.get('new_employee_cost')
        new_temporary_atd = request.POST.get('new_temporary_atd')
        prefered_vehicle = request.POST.get('prefered_vehicle')
        justification = request.POST.get('justification')
        E_plate_no = request.POST.get('E_plate_no')
        E_con_sticker = request.POST.get('E_con_sticker')
        E_model_year = request.POST.get('E_model_year')
        E_brand = request.POST.get('E_brand')
        E_make = request.POST.get('E_make')
        E_type = request.POST.get('E_type')
        approved_by = request.POST.get('approved_by')
        approved_date = request.POST.get('approved_date')
        vehicle_provider = request.POST.get('vehicle_provider')
        vehicle_plate_no = request.POST.get('vehicle_plate_no')
        vehicle_CS_no = request.POST.get('vehicle_CS_no')
        vehicle_model = request.POST.get('vehicle_model')
        vehicle_brand = request.POST.get('vehicle_brand')
        vehicle_make = request.POST.get('vehicle_make')
        vehicle_fuel_type = request.POST.get('vehicle_fuel_type')
        svv_sla = request.POST.get('svv_sla')

        saveto_service = service_vehicle(request_date=request_date, req_employee_id=req_employee_id, req_lname=req_lname, req_fname =req_fname,
            assignee_employee_id=assignee_employee_id, assignee_group=Assignee_Group, assignee_fname=assignee_fname, assignee_lname=assignee_lname,
            assignee_costcenter=assignee_costcenter, assignee_section=assignee_section, assignee_location=assignee_location, assignee_atd=assignee_atd,
            new_employee_id=new_employee_id, new_employee_fname=new_employee_fname, new_employee_lname=new_employee_lname, new_employee_cost=new_employee_cost,
            new_temporary_atd=new_temporary_atd, prefered_vehicle=prefered_vehicle, justification=justification, E_plate_no=E_plate_no, E_con_sticker=E_con_sticker,
            E_model_year =E_model_year, E_brand=E_brand, E_make=E_make, E_type=E_type, approved_by=approved_by, approved_date=approved_date,
            vehicle_provider =vehicle_provider, vehicle_plate_no=vehicle_plate_no, vehicle_CS_no=vehicle_CS_no, vehicle_model =vehicle_model,
            vehicle_brand =vehicle_brand, vehicle_make=vehicle_make, vehicle_fuel_type=vehicle_fuel_type, SVV_SLA =svv_sla,)
        saveto_service.save()

        return HttpResponseRedirect('/Request/Service/')

class serviceCreateView(SuccessMessageMixin, CreateView):
    model = service_vehicle
    form_class = serviceform
    template_name = 'service_vehicle/service_form.html'

    def get_success_message(self, cleaned_data):
        print(cleaned_data)
        return "New Service Vehicle Request Has been Created!"

class serviceUpdateView(SuccessMessageMixin, UpdateView):
    model = service_vehicle
    form_class = serviceform
    template_name = 'service_vehicle/service_form.html'
    
    def get_success_message(self, cleaned_data):
    	print(cleaned_data)
    	return "Service Vehicle Request Update Successfully!"

class serviceDetailView(DetailView):
    model = service_vehicle
    template_name = 'service_vehicle/service_details.html'

class serviceDeleteView(BSModalDeleteView):
    model = service_vehicle
    template_name = 'service_vehicle/service_delete.html'
    success_message = 'Success: Service Vehicle Request was deleted.'
    success_url = reverse_lazy('service_list')

def serviceHistoryView(request):
    if request.method == "GET":
       obj = service_vehicle.history.all()

       return render(request, 'service_vehicle/service_history.html', context={'object': obj})

def service_request_excel(request):
    service_queryset = service_vehicle.objects.all()   
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Service Vehicle Request.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Service Vehicle Request'

    columns = [
        'Request Date' ,
        'Employee Id' ,
        'Last Name' ,
        'First Name' ,
        'Assignee Employee Id' ,
        'Assignee Group' ,
        'Assignee First Name' ,
        'Assignee Last Name' ,
        'Assignee Cost Center' ,
        'Assignee Section' ,
        'Assignee Location' ,
        'Assignee ATD' ,
        'New Employee Id' ,
        'New Employee First Name' ,
        'New Employee Last Name' ,
        'New Employee Cost Center' ,
        'New Temporary ATD' ,
        'Prefered Vehicle' ,
        'Justification' ,
        'Plate No' ,
        'Conduction Sticker' ,
        'Model' ,
        'Brand' ,
        "Make" ,
        'Type' ,
        'Approved By' ,
        'Approved Date' ,
        'Vehicle Provider' ,
        'Vehicle Plate No' ,
        'Vehicle CS No' ,
        'Vehicle Model' ,
        'Vehicle Brand' ,
        'Vehicle Make' ,
        'Vehicle Fuel Type' ,
        'SLA' ,
        'Date Initiated' ,

    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for service in service_queryset:
        row_num += 1
        row = [
        service.request_date ,
        service.req_employee_id ,
        service.req_lname ,
        service.req_fname ,
        service.assignee_employee_id ,
        service.assignee_group ,
        service.assignee_fname ,
        service.assignee_lname ,
        service.assignee_costcenter ,
        service.assignee_section ,
        service.assignee_location ,
        service.assignee_atd ,
        service.new_employee_id ,
        service.new_employee_fname ,
        service.new_employee_lname ,
        service.new_employee_cost ,
        service.new_temporary_atd ,
        service.prefered_vehicle ,
        service.justification,
        service.E_plate_no ,
        service.E_con_sticker ,
        service.E_model_year ,
        service.E_brand ,
        service.E_make ,
        service.E_type ,
        service.approved_by ,
        service.approved_date ,
        service.vehicle_provider ,
        service.vehicle_plate_no ,
        service.vehicle_CS_no ,
        service.vehicle_model ,
        service.vehicle_brand ,
        service.vehicle_make ,
        service.vehicle_fuel_type ,
        service.SVV_SLA ,
        service.date_initiated ,
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response


                      #####################################  
                    #########################################
                   #####.     Vehicle Repair  Request    #####
                    #########################################
                     ######################################


class repairListView(ListView):
    model = Vehicle_Repair
    template_name='vehicle_repair/repair_list.html'
    
def repairCreate(request):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    emplist = EmployeeMasterlist.objects.all()
    vlist = VehicleMasterList.objects.all()
    return render(request, 'vehicle_repair/repair_new.html',{'Title':'Vehicle - Vehicle Repair','emplist':emplist,'vlist':vlist})

def repairsubmit(request):
    if request.method == 'POST':
        request_date = request.POST.get('request_date')
        emp_id = request.POST.get('emp_id')
        cost_center = request.POST.get('cost_center')
        fname = request.POST.get('fname')
        lname = request.POST.get('lname')
        c_no = request.POST.get('c_no')
        company = request.POST.get('company')
        department = request.POST.get('department')
        group = request.POST.get('group')
        plate_no = request.POST.get('plate_no')
        v_brand = request.POST.get('v_brand')
        engine = request.POST.get('engine')
        v_model = request.POST.get('v_model')
        v_make = request.POST.get('v_make')
        chassis = request.POST.get('chassis')
        v_band = request.POST.get('v_band')
        cs_no = request.POST.get('cs_no')
        eq_no = request.POST.get('eq_no')
        fleet_area = request.POST.get('fleet_area')
        particulars = request.POST.get('particulars')
        category = request.POST.get('category')
        maintenance_type1 = request.POST.get('maintenance_type1')
        maintenance_type2 = request.POST.get('maintenance_type2')
        scope_work1 = request.POST.get('scope_work1')
        scope_work2 = request.POST.get('scope_work2')
        recomendations = request.POST.get('recomendations')
        service_reminder = request.POST.get('service_reminder')
        repair_verified_by = request.POST.get('repair_verified_by')
        work_order1 = request.POST.get('work_order1')
        work_order2 = request.POST.get('work_order2')
        work_order3 = request.POST.get('work_order3')
        date_work_created = request.POST.get('date_work_created')
        repair_shop = request.POST.get('repair_shop')
        date_forward = request.POST.get('date_forward')
        estimate_no = request.POST.get('estimate_no')
        maintenance_amount = request.POST.get('maintenance_amount')
        less_discount = request.POST.get('less_discount')
        estimate_remark = request.POST.get('estimate_remark')
        estimate_attach = request.POST.get('estimate_attach')
        approved_by = request.POST.get('approved_by')
        kilo_reading = request.POST.get('kilo_reading')
        vrr_sla = request.POST.get('vrr_sla')

        saveto_repair = Vehicle_Repair(request_date=request_date, employee=emp_id, cost_center=cost_center, first_name=fname,
            last_name=lname, contact_no=c_no, company=company, department=department, group_section=group,
            plate_no=plate_no, v_brand=v_brand, engine=engine, v_make=v_make, v_model=v_model, chassis=chassis,
            band=v_band, cond_sticker=cs_no, equipment_no=eq_no, fleet_area=fleet_area, particulars=particulars,
            category=category, maintenance_type1=maintenance_type1, scope_work1=scope_work1, maintenance_type2=maintenance_type2,
            scope_work2=scope_work2, recommendations=recomendations, service_reminder=service_reminder, verified_by=repair_verified_by,
            work_order1=work_order1, work_order2=work_order2, work_order3=work_order3, datework_created=date_work_created,
            Shop_vendor=repair_shop, date_forwarded=date_forward, estimate_no=estimate_no, maintenance_amount=maintenance_amount,
            less_discount=less_discount, estimate_remarks=estimate_remark, estimate_attached=estimate_attach, approvedby=approved_by,
            meter_reading=kilo_reading,VRR_SLA=vrr_sla,
    )
        saveto_repair.save()

        return HttpResponseRedirect('/Request/Repair/')

class repairDetailView(DetailView):
    model = Vehicle_Repair
    template_name = 'vehicle_repair/repair_details.html'

class repairUpdateView(SuccessMessageMixin, UpdateView):
    model = Vehicle_Repair
    form_class = repairform
    template_name = 'vehicle_repair/repair_form.html'
    
    def get_success_message(self, cleaned_data):
    	print(cleaned_data)
    	return "Vehicle Repair Request Updated Successfully!"

class repairDeleteView(BSModalDeleteView):
    model = Vehicle_Repair
    template_name = 'vehicle_repair/repair_delete.html'
    success_message = 'Success: Vehicle Repair Request was deleted.'
    success_url = reverse_lazy('repair_list')

def repairHistoryView(request):
    if request.method == "GET":
       obj = Vehicle_Repair.history.all()

       return render(request, 'vehicle_repair/repair_history.html', context={'object': obj})

def repair_request_excel(request):
    repair_queryset = Vehicle_Repair.objects.all()   
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Vehicle Repair Request.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Vehicle Repair Request'

    columns = [
                'Request Date' ,
                'Employee' ,
                'Cost Center' ,
                'First Name' ,
                'Last Name' ,
                'Contact No' ,
                'Company' ,
                'Department' ,
                'Group Section' ,
                'Plate No' ,
                'Brand' ,
                'Engine' ,
                'Make' ,
                'Model' ,
                'Chassis' ,
                'Band' ,
                'Conduction Sticker' ,
                'Equipment No' ,
                'Fleet Area' ,
                'Particulars' ,
                'Category' ,
                'Maintenance Type 1' ,
                'Scope Work 1' ,
                'Maintenance Type 2' ,
                'Scope Work 2' ,
                'Recommendations' ,
                'Service Reminder' ,
                'Verified By' ,
                'Work Order 1' ,
                'Work Order 2' ,
                'Work Order 3' ,
                'Date Work Created' ,
                'Shop Vendor' ,
                'Date Forwarded' ,
                'Estimate No' ,
                'Maintenance Amount' ,
                'Less Discount' ,
                'Estimate Remarks' ,
                'Estimate Attached' ,
                'Approved By' ,
                'Meter Reading' ,
                'SLA' ,
                'Date Initiated' ,

    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for repair in repair_queryset:
        row_num += 1
        row = [
        request_date ,
                repair.employee ,
                repair.cost_center ,
                repair.first_name ,
                repair.last_name ,
                repair.contact_no ,
                repair.company ,
                repair.department ,
                repair.group_section ,
                repair.plate_no ,
                repair.v_brand ,
                repair.engine ,
                repair.v_make ,
                repair.v_model ,
                repair.chassis ,
                repair.band ,
                repair.cond_sticker ,
                repair.equipment_no ,
                repair.fleet_area ,
                repair.particulars ,
                repair.category ,
                repair.maintenance_type1 ,
                repair.scope_work1 ,
                repair.maintenance_type2 ,
                repair.scope_work2 ,
                repair.recommendations ,
                repair.service_reminder ,
                repair.verified_by ,
                repair.work_order1 ,
                repair.work_order2 ,
                repair.work_order3 ,
                repair.datework_created ,
                repair.Shop_vendor ,
                repair.date_forwarded ,
                repair.estimate_no ,
                repair.maintenance_amount ,
                repair.less_discount ,
                repair.estimate_remarks ,
                repair.estimate_attached ,
                repair.approvedby ,
                repair.meter_reading ,
                repair.VRR_SLA ,
                repair.date_initiated ,
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response

