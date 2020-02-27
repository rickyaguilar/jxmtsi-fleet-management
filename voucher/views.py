from django.shortcuts import render,HttpResponseRedirect,HttpResponse
from django.contrib.messages.views import SuccessMessageMixin
from django.views import generic
from openpyxl import Workbook
from django.urls import reverse_lazy
from .models import (
   	expense_voucher,
)
from masterlist.models import EmployeeMasterlist,VehicleMasterList
from . forms import (
    voucherform,
)
from django.views.generic import (
                				CreateView,
                				ListView,
                				UpdateView,
                				DetailView,
                				)
from bootstrap_modal_forms.generic import (
                                           BSModalDeleteView
                                           )

def voucher(request):
	employee_list = EmployeeMasterlist.objects.all()
	vehicle_list = VehicleMasterList.objects.all()
	return render(request, 'voucher_new.html', {'title': 'Voucher - Create New Expense Voucher', 'employee_list': employee_list, 
	'vehicle_list': vehicle_list})

class voucherListView(ListView):
	model = expense_voucher
	template_name = 'voucher_list.html'

def vouchersubmit(request):
    if request.method == 'POST':
        employee_id = request.POST.get('employee_id')
        employee_fname = request.POST.get('employee_fname')
        employee_lname = request.POST.get('employee_lname')
        employee_group = request.POST.get('employee_group')
        employee_cost = request.POST.get('employee_cost')
        new_employee_id = request.POST.get('new_employee_id')
        new_employee_fname = request.POST.get('new_employee_fname')
        new_employee_lname = request.POST.get('new_employee_lname')
        new_employee_group = request.POST.get('new_employee_group')
        new_employee_cost = request.POST.get('new_employee_cost')
        fleet_area = request.POST.get('fleet_area')
        date_recieved_voucher = request.POST.get('date_recieved_voucher')
        trans_type = request.POST.get('trans_type')
        voucher_no = request.POST.get('voucher_no')
        voucher_amount = request.POST.get('voucher_amount')
        voucher_type = request.POST.get('voucher_type')
        fuel_amount = request.POST.get('fuel_amount')
        fuel_products = request.POST.get('fuel_products')
        fuel_liters = request.POST.get('fuel_liters')
        service_amount = request.POST.get('service_amount')
        workorder_no = request.POST.get('workorder_no')
        plate_no = request.POST.get('plate_no')
        odometer_start = request.POST.get('odometer_start')
        odometer_end = request.POST.get('odometer_end')
        v_brand = request.POST.get('v_brand')
        v_make = request.POST.get('v_make')
        v_fuel_type = request.POST.get('v_fuel_type')
        v_model = request.POST.get('v_model')
        new_plate_no = request.POST.get('new_plate_no')
        new_odometer_start = request.POST.get('new_odometer_start')
        new_odometer_end = request.POST.get('new_odometer_end')
        new_v_brand = request.POST.get('new_v_brand')
        new_v_make = request.POST.get('new_v_make')
        new_vfuel_type = request.POST.get('new_vfuel_type')
        new_vmodel = request.POST.get('new_vmodel')
        gt_admin = request.POST.get('gt_admin')
        approval_date = request.POST.get('approval_date')
        imm_supervisor = request.POST.get('imm_supervisor')
        im_approval_date = request.POST.get('im_approval_date')
        voucher_docs1 = request.POST.get('voucher_docs1')
        voucher_docs2 = request.POST.get('voucher_docs2')
        voucher_docs3 = request.POST.get('voucher_docs3')
        voucher_remarks = request.POST.get('voucher_remarks')
        date_forward_voucher = request.POST.get('date_forward_voucher')
        evo_sla = request.POST.get('evo_sla')

        saveto_voucher = expense_voucher(employee_id=employee_id, employee_fname=employee_fname, employee_lname=employee_lname, 
        	employee_group=employee_group, employee_cost=employee_cost, new_employee_id=new_employee_id, 
        	new_employee_fname=new_employee_fname, new_employee_lname=new_employee_lname, new_employee_group=new_employee_group, 
        	new_employee_cost=new_employee_cost, fleet_area=fleet_area, received_voucher=date_recieved_voucher, 
        	trans_type=trans_type, voucher_no=voucher_no, voucher_amount=voucher_amount, voucher_type=voucher_type, 
        	fuel_amount=fuel_amount, fuel_products=fuel_products, fuel_liters=fuel_liters, service_amount=service_amount, 
        	work_order=workorder_no, plate_number=plate_no, odometer_start=odometer_start, odometer_end=odometer_end, 
        	v_brand=v_brand, v_make=v_make, v_fuel_type=v_fuel_type, v_model=v_model, new_plate_number=new_plate_no, 
        	new_odometer_start=new_odometer_start, new_odometer_end=new_odometer_end, new_v_brand=new_v_brand, 
        	new_v_make=new_v_make, new_v_fuel_type=new_vfuel_type, new_v_model=new_vmodel, gt_admin=gt_admin, 
        	approval_date=approval_date, immediate_supervisor=imm_supervisor, im_approval_date=im_approval_date, 
        	voucher_docs1=voucher_docs1, voucher_docs2=voucher_docs2, voucher_docs3=voucher_docs3, voucher_remarks=voucher_remarks, 
        	voucher_forwarded=date_forward_voucher, EVO_SLA=evo_sla,)
        saveto_voucher.save()

        return HttpResponseRedirect('/Voucher/Voucher/')

class voucherUpdate(SuccessMessageMixin, UpdateView):
	model = expense_voucher
	form_class = voucherform
	template_name = 'voucher_form.html'

	def get_success_message(self, cleaned_data):
		print(cleaned_data)
		return "Expense Voucher Updated Successfully!"
		
class voucherDetails(DetailView):
	model = expense_voucher
	template_name = 'voucher_details.html'
	
class voucherDeleteView(BSModalDeleteView):
    model = expense_voucher
    template_name = 'voucher_delete.html'
    success_message = 'Success: Expense Voucher was deleted.'
    success_url = reverse_lazy('voucher_list')

def voucherHistoryView(request):
    if request.method == "GET":
       obj = expense_voucher.history.all()

       return render(request, 'voucher_history.html', context={'object': obj})

def voucher_excel(request):
    voucher_queryset = expense_voucher.objects.all()   
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename= Expense Voucher.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Expense Voucher'

    columns = [
    'Employee Id' ,
    'Employee Fist Name' ,
    'Employee Last Name' ,
    'Employee Group' ,
    'Employee Cost Center' ,
    'New Employee Id' ,
    'New Employee First Name' ,
    'New Employee Last Name' ,
    'New Employee Group' ,
    'New Employee Cost Center' ,
    'Fleet Area' ,
    'Received Voucher' ,
    'Transaction Type' ,
    'Voucher No' ,
    'Voucher Amount' ,
    'Voucher Type' ,
    'Fuel Amount' ,
    'Fuel Products' ,
    'Fuel Liters' ,
    'Service Amount' ,
    'Work Order' ,
    'Plate Number' ,
    'Odometer Start' ,
    'Odometer End' ,
    'Brand' ,
    ' Make' ,
    'Fuel Type' ,
    'Model' ,
    'New Plate Number' ,
    'New Odometer Start' ,
    'New Odometer End' ,
    'New Brand' ,
    'New Make' ,
    'New Fuel Type' ,
    'New Model' ,
    'Gt Admin' ,
    'Approval Date' ,
    'Immediate Supervisor' ,
    'Approval Date' ,
    'Voucher Documents 1' ,
    'Voucher Documents 2' ,
    'Voucher Documents 3' ,
    'Voucher Remarks' ,
    'Voucher Forwarded' ,
    'SLA' ,
    'Date Initiated' ,

    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for voucher in voucher_queryset:
        row_num += 1
        row = [
            voucher.employee_id ,
            voucher.employee_fname ,
            voucher.employee_lname ,
            voucher.employee_group ,
            voucher.employee_cost ,
            voucher.new_employee_id ,
            voucher.new_employee_fname ,
            voucher.new_employee_lname ,
            voucher.new_employee_group ,
            voucher.new_employee_cost ,
            voucher.fleet_area ,
            voucher.received_voucher ,
            voucher.trans_type ,
            voucher.voucher_no ,
            voucher.voucher_amount ,
            voucher.voucher_type ,
            voucher.fuel_amount ,
            voucher.fuel_products ,
            voucher.fuel_liters ,
            voucher.service_amount ,
            voucher.work_order ,
            voucher.plate_number ,
            voucher.odometer_start ,
            voucher.odometer_end ,
            voucher.v_brand ,
            voucher.v_make ,
            voucher.v_fuel_type ,
            voucher.v_model ,
            voucher.new_plate_number ,
            voucher.new_odometer_start ,
            voucher.new_odometer_end ,
            voucher.new_v_brand ,
            voucher.new_v_make ,
            voucher.new_v_fuel_type ,
            voucher.new_v_model ,
            voucher.gt_admin ,
            voucher.approval_date ,
            voucher.immediate_supervisor ,
            voucher.im_approval_date ,
            voucher.voucher_docs1 ,
            voucher.voucher_docs2 ,
            voucher.voucher_docs3 ,
            voucher.voucher_remarks ,
            voucher.voucher_forwarded ,
            voucher.EVO_SLA ,
            voucher.date_initiated ,
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response



        