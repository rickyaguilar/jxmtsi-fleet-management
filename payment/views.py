from django.shortcuts import render,HttpResponseRedirect,reverse,HttpResponse
from django.contrib.messages.views import SuccessMessageMixin
from openpyxl import Workbook
from django.urls import reverse_lazy
from django.views import generic
import datetime
from datetime import date
from .models import (
    CarRental,
    VehiclePayment,
    Fuel_supplier,
    Vehicle_Repair_payment,
)
from masterlist.models import EmployeeMasterlist,VehicleMasterList
from django.views.generic import (
     DetailView,
     ListView,
     CreateView,
     UpdateView,
 )
from . forms import (
    VehiclePaymentform,
    FuelsupplierForm,
    vrepair_form
)
from bootstrap_modal_forms.generic import (
                                           BSModalDeleteView
                                           )

class CarListView(ListView):
	model = CarRental
	template_name = 'payment/car/carrental_list.html'
	
class CarRentalDetailView(DetailView):
	model = CarRental
	template_name = 'payment/car/carrental_summary.html'

def Carrentalpayment(request):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    e_list = EmployeeMasterlist.objects.all()
    v_list = VehicleMasterList.objects.all()
    return render(request, 'payment/car/car_rental1.html',{'title': 'Car - Car Rental', 'e_list': e_list, 'v_list': v_list})


def Carrental_submit(request):
    if request.method == 'POST':
    #<---Assign Details--->
        Bill_date = request.POST.get('Bdate')
        employee_id = request.POST.get('Eid')
        firstname = request.POST.get('Efm')
        lastname = request.POST.get('Elm')
        Ass_company = request.POST.get('Acom')
        Cost_center = request.POST.get('Acost')
        Date_initiated = request.POST.get('date')
        #<-- other assign-->
        Other_assigneeFM = request.POST.get('Ofname')
        Other_assigneeLM = request.POST.get('Olname')
        Other_cost = request.POST.get('Ocost')
        #<--Vehicle Details-->
        Plate_no = request.POST.get('Pnumber')
        V_brand = request.POST.get('Vbrand')
        V_model = request.POST.get('Vmodel')
        V_make = request.POST.get('Vmake')
        #<--Rental Details-->
        Delivered_V = request.POST.get('Ddelivered')
        S_rental = request.POST.get('Srental')
        E_rental = request.POST.get('Erental')
        #<--Expense Details-->
        R_cost = request.POST.get('Rcost')
        Gas_cost = request.POST.get('Gcost')
        Toll_fee = request.POST.get('Tfee')
        Park_fee = request.POST.get('Pfee')
        Del_fee = request.POST.get('Dfee')
        Driverfee = request.POST.get('Driverfee')
        Meal_cost = request.POST.get('M_cost')
        Other_exp = request.POST.get('Other_exp')
        Total = request.POST.get('Total')
        C_SLA = request.POST.get('SLA')
        car_provider = request.POST.get('car_provider')
        sqa_number = request.POST.get('sqa_number')
        rfp_no2 = request.POST.get('rfp_no2')

		#############################
		######Date calculator########
		#############################

        d1 = datetime.datetime.strptime(S_rental, '%Y-%m-%d')
        d2 = datetime.datetime.strptime(E_rental, '%Y-%m-%d')
        Rduration = (d2 - d1)
		
		

        saveto_CRP = CarRental(Bill_date=Bill_date, Employee_id=employee_id, L_name=lastname, F_name=firstname,
            Assignee_company=Ass_company, Cost_center=Cost_center, O_Fname=Other_assigneeFM, O_Lname=Other_assigneeLM,
            O_cost_center=Other_cost, Plate_no=Plate_no, V_brand=V_brand, V_make=V_make,
            D_vehicle=Delivered_V, S_rental=S_rental, E_rental=E_rental, R_duration=Rduration, R_Cost=R_cost,
            G_cost=Gas_cost, T_fee=Toll_fee, P_fee=Park_fee, Del_fee=Del_fee, Dri_fee=Driverfee, M_cost=Meal_cost,
            O_expenses=Other_exp, T_expenses=Total, Date_initiated=Date_initiated, C_SLA=C_SLA, car_provider = car_provider, sqa_number = sqa_number,
            rfp_no2 = rfp_no2)
        saveto_CRP.save()

        return HttpResponseRedirect('/Payment/Car/')

def Carrental_update(request, pk):
    if request.method == 'POST':
    #<---Assign Details--->
        Bill_date = request.POST.get('Bdate')
        employee_id = request.POST.get('Eid')
        firstname = request.POST.get('Efm')
        lastname = request.POST.get('Elm')
        Ass_company = request.POST.get('Acom')
        Cost_center = request.POST.get('Acost')
        #<-- other assign-->
        Other_assigneeFM = request.POST.get('Ofname')
        Other_assigneeLM = request.POST.get('Olname')
        Other_cost = request.POST.get('Ocost')
        #<--Vehicle Details-->
        Plate_no = request.POST.get('Pnumber')
        V_brand = request.POST.get('Vbrand')
        V_make = request.POST.get('Vmake')
        #<--Rental Details-->
        Delivered_V = request.POST.get('Ddelivered')
        S_rental = request.POST.get('Startrental')
        E_rental = request.POST.get('Endrental')
        #<--Expense Details-->
        R_cost = request.POST.get('Rentcost')
        Gas_cost = request.POST.get('Gascost')
        Toll_fee = request.POST.get('Tollfee')
        Park_fee = request.POST.get('Parkingfee')
        Del_fee = request.POST.get('Delfee')
        Driverfee = request.POST.get('Drifee')
        Meal_cost = request.POST.get('Mealcost')
        Other_exp = request.POST.get('Otherexp')
        Total = request.POST.get('TotalExp')
        car_provider = request.POST.get('car_provider')
        sqa_number = request.POST.get('sqa_number')
        rfp_no2 = request.POST.get('rfp_no2')

		#############################
		######Date calculator########
		#############################

        d1 = datetime.datetime.strptime(S_rental, '%Y-%m-%d').date()
        d2 = datetime.datetime.strptime(E_rental, '%Y-%m-%d').date()
        Rduration = (d2 - d1)

        CarRental.objects.filter(id=pk).update(Bill_date=Bill_date,Employee_id=employee_id, L_name=lastname, F_name=firstname,
            Assignee_company=Ass_company, Cost_center=Cost_center, O_Fname=Other_assigneeFM, O_Lname=Other_assigneeLM,
            O_cost_center=Other_cost, Plate_no=Plate_no, V_brand=V_brand, V_make=V_make,
            D_vehicle=Delivered_V, S_rental=S_rental, E_rental=E_rental, R_duration=Rduration, R_Cost=R_cost,
            G_cost=Gas_cost, T_fee=Toll_fee, P_fee=Park_fee, Del_fee=Del_fee, Dri_fee=Driverfee, M_cost=Meal_cost,
            O_expenses=Other_exp, T_expenses=Total)

        return HttpResponseRedirect('/Payment/Car/')

class carrentalDeleteView(BSModalDeleteView):
    model = CarRental
    template_name = 'payment/carrental_delete.html'
    success_message = 'Success: Item was deleted.'
    success_url = reverse_lazy('carrental_list')

def carrentalHistoryView(request):
    if request.method == "GET":
       obj = CarRental.history.all()

       return render(request, 'payment/car/carrental_history.html', context={'object': obj})

def rent(request):
	emp = EmployeeMasterlist.objects.all()
	vechicle = VehicleMasterList.objects.all()
	return render(request, 'payment/car_rental.html', {'title': 'CarRental - Create New Car Rental Request', 'emp': emp, 
	'vechicle': vechicle})

def vehiclecreate(request):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    elist = EmployeeMasterlist.objects.all()
    vlist = VehicleMasterList.objects.all()
    return render(request, 'payment/vehicle/vehiclepayment_form.html',{'elist':elist,'vlist':vlist})

def vehicle_submit(request):
    if request.method == 'POST':
        a_emp_id = request.POST.get('a_emp_id')
        emp_fname = request.POST.get('emp_fname')
        emp_lname = request.POST.get('emp_lname')
        Delivery_Date = request.POST.get('Delivery_Date')
        Plate_Number = request.POST.get('Plate_Number')
        Model_Year = request.POST.get('Model_Year')
        Brand = request.POST.get('Brand')
        Make = request.POST.get('Make')
        Dealer = request.POST.get('Dealer')
        date_received = request.POST.get('Date_Received_LTO_Documents')
        Docs_plate_no = request.POST.get('Docs_plate_no')
        LTO_stickers = request.POST.get('LTO_stickers')
        Sticker_fields = request.POST.get('Sticker_fields')
        Date_initial = request.POST.get('Date_initial')
        First_payment = request.POST.get('First_payment')
        LTO_Charges = request.POST.get('LTO_Charges')
        Outstanding_Balance = request.POST.get('Outstanding_Balance')
        Date_final = request.POST.get('Date_final')
        Routing_Remarks = request.POST.get('Routing_Remarks')
        v_sla = request.POST.get('v_sla')
        rfp_number = request.POST.get('rfp')
        invoice_number = request.POST.get('invno')
        equip_no = request.POST.get('equip_no')
        asset_no = request.POST.get('asset_no')
        sap_no = request.POST.get('sap_no')
        mat_no = request.POST.get('mat_no')
        Dealer_name = request.POST.get('Dealer_name')
        
        saveto_v = VehiclePayment(A_employee_ID = a_emp_id,E_First_name = emp_fname,E_Last_name = emp_lname,V_deliverDate = Delivery_Date,
        Plate_no = Plate_Number,V_model = Model_Year,V_brand = Brand,V_make = Make,V_dealer = Dealer,LTO_documents = date_received,
        Docs_plate_no = Docs_plate_no,LTO_stickers = LTO_stickers,Sticker_fields = Sticker_fields,Date_initial = Date_initial,
        First_payment = First_payment,LTO_charges = LTO_Charges,Outstanding_balance = Outstanding_Balance,Date_final = Date_final,
        Routing_remarks = Routing_Remarks,V_SLA = v_sla,rfp_number = rfp_number,invoice_number = invoice_number, equip_no = equip_no,
        asset_no = asset_no, sap_no = sap_no, mat_no = mat_no, Dealer_name = Dealer_name)
        saveto_v.save()

        return HttpResponseRedirect('/Payment/Vehicle/')

class VehicleListView(ListView):
	model = VehiclePayment
	template_name = 'payment/vehicle/vehicle_list.html'

class VehicleDetailView(DetailView):
	model = VehiclePayment
	template_name = 'payment/vehicle/vehiclepayment_summary.html'

class VehicleUpdateView(SuccessMessageMixin, UpdateView):
    model = VehiclePayment
    form_class = VehiclePaymentform
    template_name = 'payment/vehicle/vehiclepayment_new.html'

    def get_success_message(self, cleaned_data):
    	print(cleaned_data)
    	return "New Vehicle Payment's Update Successfully!"

class VehicleDeleteView(BSModalDeleteView):
    model = VehiclePayment
    template_name = 'vehicle/vehicle_delete.html'
    success_message = 'Success: Item was deleted.'
    success_url = reverse_lazy('Vehicle_list')

def VehicleHistoryView(request):
    if request.method == "GET":
       obj = VehiclePayment.history.all()

       return render(request, 'payment/vehicle/vehicle_history.html', context={'object': obj})
			
class FuelDetailView(DetailView):
	model = Fuel_supplier
	template_name = 'payment/fuel/fuel_supplierSummary.html'

class FuelListView(ListView):
	model = Fuel_supplier
	template_name = 'payment/fuel/fuel_supplierList.html'

class FuelCreateView(SuccessMessageMixin, CreateView):
    model = Fuel_supplier
    form_class = FuelsupplierForm
    template_name = 'payment/fuel/fuel_supplier.html'

    def get_success_message(self, cleaned_data):
    	print(cleaned_data)
    	return "Fuel Supplier Has been Created!"

class FuelUpdateView(SuccessMessageMixin, UpdateView):
    model = Fuel_supplier
    form_class = FuelsupplierForm
    template_name = 'payment/fuel/fuel_supplier.html'

    def get_success_message(self, cleaned_data):
    	print(cleaned_data)
    	return "Fuel Supplier Update Successfully!"

class FuelDeleteView(BSModalDeleteView):
    model = Fuel_supplier
    template_name = 'fuel/fuel_supplier_delete.html'
    success_message = 'Success: Item was deleted.'
    success_url = reverse_lazy('Fuel_supplierList')

def FuelHistoryView(request):
    if request.method == "GET":
       obj = Fuel_supplier.history.all()

       return render(request, 'payment/fuel/fuel_supplier_history.html', context={'object': obj})

class vrepair_payment(ListView):
    model = Vehicle_Repair_payment
    template_name = 'payment/vehicle_repair/vehicle_repair_paymentList.html'

def vrepair_payment_create(request):
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    emplist = EmployeeMasterlist.objects.all()
    vlist = VehicleMasterList.objects.all()
    return render(request, 'payment/vehicle_repair/vehicle_repair_form.html',{'emplist':emplist,'vlist':vlist})

class vrepairDetailView(DetailView):
    model = Vehicle_Repair_payment
    template_name = 'payment/vehicle_repair/vehicle_repair_details.html'

class vrepairDeleteView(BSModalDeleteView):
    model = Vehicle_Repair_payment
    template_name = 'vehicle_repair/vehicle_repair_delete.html'
    success_message = 'Success: Item was deleted.'
    success_url = reverse_lazy('vehiclerepair_payment')

class vrepairUpdate(SuccessMessageMixin, UpdateView):
    model = Vehicle_Repair_payment
    form_class = vrepair_form
    template_name = 'payment/vehicle_repair/vehicle_repair_update.html'

    def get_success_message(self, cleaned_data):
        print(cleaned_data)
        return "Vehicle Repair Payment Update Successfully!"

def vrepairlHistoryView(request):
    if request.method == "GET":
       obj = Vehicle_Repair_payment.history.all()

       return render(request, 'payment/vehicle_repair/vehicle_repair_history.html', context={'object': obj})

def vrepairsubmit(request):
    if request.method == 'POST':
        request_date = request.POST.get('request_date')
        emp_id = request.POST.get('emp_id')
        cost_center = request.POST.get('cost_center')
        fname = request.POST.get('fname')
        lname = request.POST.get('lname')
        c_no = request.POST.get('c_no')
        company = request.POST.get('company')
        department = request.POST.get('department')
        group = request.POST.get('group')
        plate_no = request.POST.get('plate_no')
        v_brand = request.POST.get('v_brand')
        engine = request.POST.get('engine')
        v_model = request.POST.get('v_model')
        v_make = request.POST.get('v_make')
        chassis = request.POST.get('chassis')
        v_band = request.POST.get('v_band')
        cs_no = request.POST.get('cs_no')
        eq_no = request.POST.get('eq_no')
        dealership = request.POST.get('dealer')
        amount = request.POST.get('amount')
        service_type = request.POST.get('service_type')
        rfp_no = request.POST.get('rfp_no')
        invoice_number2 = request.POST.get('invoice_num')
        invoice_date = request.POST.get('invoice_date')


        saveto_vrp = Vehicle_Repair_payment(request_date=request_date, employee=emp_id, cost_center=cost_center, first_name=fname,
            last_name=lname, contact_no=c_no, company=company, department=department, group_section=group,
            plate_no=plate_no, v_brand=v_brand, engine=engine, v_make=v_make, v_model=v_model, chassis=chassis,
            band=v_band, cond_sticker=cs_no, equipment_no=eq_no, dealership=dealership, amount=amount, service_type=service_type, rfp_no =rfp_no, invoice_number2=invoice_number2,
            invoice_date=invoice_date
    )
        saveto_vrp.save()

        return HttpResponseRedirect('/Payment/VehicleRepair/')


def car_excel(request):
    car_queryset = CarRental.objects.all()   
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Car Rental Payment.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Car Rental Payment'

    columns = [
            'Bill date',
            'Employee Id',
            'Last Name',
            'First Name',
            'Company',
            'Cost Center',
            'Date Created',
            'Other Assignee First Name',
            'Other Assignee Last Name',
            'Other Cost Center',
            'Plate No',
            'Model',
            'Brand',
            'Make',
            'Date Rental',
            'Start Date Rental',
            'End Date Rental',
            'Rent Duration',
            'Rent Cost',
            'Gasoline Cost',
            'Toll Fee',
            'Parking Fee',
            'Delivery Fee',
            'Driver Fee',
            'Meal Cost',
            'Other Expense',
            'Total Expense',
            'car_provider',
            'sqa_number',
            'rfp_no2',
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for car in car_queryset:
        row_num += 1
        # ordate = car.ORIGINAL_OR_DATE.strftime('%m/%d/%Y')
        # platerelease = car.PLATE_NUMBER_RELEASE_DATE.strftime('%m/%d/%Y')
        row = [
                car.Bill_date,
                car.Employee_id,
                car.L_name,
                car.F_name,
                car.Assignee_company,
                car.Cost_center,
                car.Date_initiated,
                car.O_Fname,
                car.O_Lname,
                car.O_cost_center,
                car.Plate_no,
                car.V_model,
                car.V_brand,
                car.V_make,
                car.D_vehicle,
                car.S_rental,
                car.E_rental,
                car.R_duration,
                car.R_Cost,
                car.G_cost,
                car.T_fee,
                car.P_fee,
                car.Del_fee,
                car.Dri_fee,
                car.M_cost,
                car.O_expenses,
                car.T_expenses,
                car.sqa_number,
                car.car_provider,
                car.rfp_no2,
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response

def vehicle_excel(request):
    vehicle_queryset = VehiclePayment.objects.all()   
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Vehicle Payment.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Vehicle Payment'

    columns = [
            'PO Number',
            'Employee Id',
            'First Name',
            'Last Name',
            'Delivery Date',
            'Plate No',
            'Model',
            'Brand',
            'Make',
            'Dealer',
            'LTO Documents',
            'Documents Plate No',
            'LTO Conduction Stickers',
            'Date Initial',
            'First Payment',
            'LTO Charges',
            'Outstanding Balance',
            'Date Final',
            'Remarks',
            'Date Initiated',
            'rfp_number',
            'invoice_number',
            'equip_no',
            'asset_no',
            'sap_no',
            'mat_no',
            'Dealer_name',
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for vehicle in vehicle_queryset:
        row_num += 1
        # ordate = car.ORIGINAL_OR_DATE.strftime('%m/%d/%Y')
        # platerelease = car.PLATE_NUMBER_RELEASE_DATE.strftime('%m/%d/%Y')
        row = [
                vehicle.PO_no,
                vehicle.A_employee_ID,
                vehicle.E_First_name,
                vehicle.E_Last_name,
                vehicle.V_deliverDate,
                vehicle.Plate_no,
                vehicle.V_model,
                vehicle.V_brand,
                vehicle.V_make,
                vehicle.V_dealer,
                vehicle.LTO_documents,
                vehicle.Docs_plate_no,
                vehicle.LTO_stickers,
                vehicle.Date_initial,
                vehicle.First_payment,
                vehicle.LTO_charges,
                vehicle.Outstanding_balance,
                vehicle.Date_final,
                vehicle.Routing_remarks,
                vehicle.Date_initiated,
                vehicle.rfp_number,
                vehicle.invoice_number,
                vehicle.equip_no,
                vehicle.asset_no,
                vehicle.sap_no,
                vehicle.mat_no,
                vehicle.Dealer_name,
                
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response

def fuel_excel(request):
    fuel_queryset = Fuel_supplier.objects.all()   
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Fuel Supplier Payment.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Fuel Supplier Payment'

    columns = [
    'Date Received',
	'Fuel Provider',
	'Bill Date',
	'Current Amount',
	'Outstanding Amount',
	'Payee',
	'Attached',
	'Payment Deadline',
	'Date Forwarded',
	'Date Initiated',
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for fuel in fuel_queryset:
        row_num += 1
        # ordate = car.ORIGINAL_OR_DATE.strftime('%m/%d/%Y')
        # platerelease = car.PLATE_NUMBER_RELEASE_DATE.strftime('%m/%d/%Y')
        row = [
				fuel.SOA_Date_received,
				fuel.Fuel_provider,
				fuel.SOA_billdate,
				fuel.SOA_current_amount,
				fuel.SOA_outstanding_amount,
				fuel.Payee,
				fuel.SOA_attached,
				fuel.Payment_deadline,
				fuel.Date_forwarded,
				fuel.Date_initiated,
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response
def vrepair_excel(request):
    vrepair_queryset = Vehicle_Repair_payment.objects.all()   
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Vehicle Repair Payment.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Vehicle Repair Payment'

    columns = [
            'Activity Id', 
            'Request Date', 
            'Employee', 
            'Cost Center', 
            'First Name', 
            'Last Name', 
            'Contact No', 
            'Company', 
            'Department', 
            'Group Section', 
            'Plate No', 
            'Brand', 
            'Engine', 
            'Make', 
            'Model', 
            'Chassis', 
            'Band', 
            'Conduction Sticker', 
            'Equipment No', 
            'Dealership', 
            'Amount', 
            'Service Type', 
            'Date Initiated',
            'rfpno',
            'invoice_number2',
            'invoice_date', 
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for vrp in vrepair_queryset:
        row_num += 1
        # ordate = car.ORIGINAL_OR_DATE.strftime('%m/%d/%Y')
        # platerelease = car.PLATE_NUMBER_RELEASE_DATE.strftime('%m/%d/%Y')
        row = [
            vrp.Activity_id, 
            vrp.request_date, 
            vrp.employee, 
            vrp.cost_center, 
            vrp.first_name, 
            vrp.last_name, 
            vrp.contact_no, 
            vrp.company, 
            vrp.department, 
            vrp.group_section, 
            vrp.plate_no, 
            vrp.v_brand, 
            vrp.engine, 
            vrp.v_make, 
            vrp.v_model, 
            vrp.chassis, 
            vrp.band, 
            vrp.cond_sticker, 
            vrp.equipment_no, 
            vrp.dealership, 
            vrp.amount, 
            vrp.service_type, 
            vrp.date_initiated,
            vrp.rfp_no,
            vrp.invoice_number2,
            vrp.invoice_date, 
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response





